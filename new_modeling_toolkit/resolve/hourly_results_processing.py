import datetime
import glob
import pathlib
import re
from typing import Iterable
from typing import Literal
from typing import Optional
from typing import Tuple

import numpy as np
import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from joblib import delayed
from joblib import Parallel
from loguru import logger
from openpyxl.utils import range_boundaries
from plotly.graph_objs import Figure
from plotly.subplots import make_subplots
from tqdm import tqdm

from new_modeling_toolkit.core.component import Component
from new_modeling_toolkit.system.electric.resources import GenericResource

# Excel named ranges:
HOURLY_AGG_SETTINGS_NAMED_RANGE = "hourly_aggregation_settings"
COLOR_SETTINGS_NAMED_RANGE = "build_group_colors"

# Data Column Constants
ANNUAL_POWER_INPUT_VAR_COST = "Annual Power Input Variable Cost ($)"
ANNUAL_POWER_OUTPUT_VAR_COST = "Annual Power Output Variable Cost ($)"
CHARGE_EFF = "Charging Efficiency"
CHRONO_PERIOD = "chrono_period"
CHRONO_PERIODS = "CHRONO_PERIODS"
CHRONO_TIMESTAMP = "chrono_timestamp"
COMPONENT = "Component"
COMPONENT_NAME = "Component Name"
COMPONENTTYPE = "ComponentType"
COMPONENT_TYPE = "Component Type"
CURTAILED = "Curtailed Energy (MWh)"
CURTAILMENT = "Curtailment (MW)"
DISCHG_EFF = "Discharging Efficiency"
DISPATCH_TIMESTAMP = "dispatch_timestamp"
DISPATCH_WINDOW = "dispatch_window"
DISPATCH_WINDOW_WEIGHT = "weight"
EXPORTS = "Exports"
FUELS = "Fuel(s)"
IMPORTS = "Imports"
INPUT_LOAD = "Input Load (MW)"
HOURLY_LOAD_MWH = "Hourly Load (MWh)"
LOAD_INC_CHG = "Load Including Charging"
TOTAL_LOAD_PLUS_EXPORTS = "Total Load Plus Exports"
TOTAL_LOAD_PLUS_OVERGEN = "Total Load Plus Overgen"
MODELED_YEARS = "MODELED_YEARS"
MODELEDYEAR = "modeled_year"
MODELED_YEAR = "Modeled Year"
NET_TX_PWR_MW = "Net Transmit Power (MW)"
NET_IMPORTS = "Net Imports"
OPERATIONAL_CAPACITY = "Operational Capacity (MW)"
OPERATIONAL_STORAGE_CAPACITY = "Operational Storage Capacity (MWh)"
OPN_GROUP = "Operational Group"
OVERGEN = "Hourly Overgeneration (MW)"
POWER_IN_MW = "Power Input (MW)"
POWER_INPUT_VAR_COST = "Power Input Variable Cost ($)"
POWER_NET_MW = "Net Power Output (MW)"
POWER_OUT_MW = "Power Output (MW)"
POWER_OUTPUT_VAR_COST = "Power Output Variable Cost ($)"
PRICES = "Hourly Energy Price Unweighted ($/MWh)"
REP_DAY_INT = "rep_day_as_int"
SOC_IIJ_MWh = "SOC Inter-Intra Joint (MWh)"
SOC_INTRA_MWh = "SOC Intra Period (MWh)"
TIMESTAMP = "timestamp"
TIMESTAMPS = "TIMESTAMPS"
UNSERVED_ENERGY = "Hourly Unserved Energy (MW)"
WEATHER_YEAR = "'weather' date"
WEATHERYEAR_IDX_PROFILE = "Weather Year-indexed Profile"
ZONES = "Zone(s)"
ZONAL_GROSS_IMPORTS = "Zonal Gross Imports (MW)"
ZONAL_GROSS_EXPORTS = "Zonal Gross Exports (MW)"
ZONAL_NET_IMPORTS = "Zonal Net Imports (MW)"
ZONAL_TOTAL_LOAD = "Total Load (MW)"

# Operational Columns
OP_COLS_SUPERSET = [
    OPERATIONAL_CAPACITY,
    OPERATIONAL_STORAGE_CAPACITY,
    CURTAILED,
    CURTAILMENT,
    POWER_OUT_MW,
    POWER_IN_MW,
    INPUT_LOAD,
    ZONAL_GROSS_IMPORTS,
    ZONAL_GROSS_EXPORTS,
    ZONAL_NET_IMPORTS,
    OVERGEN,
    UNSERVED_ENERGY,
    PRICES,
    NET_TX_PWR_MW,
    POWER_NET_MW,
    SOC_IIJ_MWh,
    SOC_INTRA_MWh,
]

HOURLY_OP_COLS = [
    OPN_GROUP,
    CURTAILED,
    CURTAILMENT,
    POWER_OUT_MW,
    POWER_IN_MW,
    INPUT_LOAD,
    ZONAL_GROSS_IMPORTS,
    ZONAL_GROSS_EXPORTS,
    ZONAL_NET_IMPORTS,
    OVERGEN,
    UNSERVED_ENERGY,
    PRICES,
    POWER_NET_MW,
    CHARGE_EFF,
    DISCHG_EFF,
    HOURLY_LOAD_MWH,
    WEATHERYEAR_IDX_PROFILE,
    NET_TX_PWR_MW,
    SOC_INTRA_MWh,
]

CHRONO_OP_COLS = [
    SOC_IIJ_MWh,
    CHRONO_PERIOD,
]

ADDITIONAL_COLUMNS = [
    TIMESTAMP,
    MODELEDYEAR,
    CHRONO_PERIOD,
    DISPATCH_WINDOW,
    TIMESTAMPS,
    CHRONO_PERIODS,
    MODELED_YEARS,
    MODELED_YEAR,
    INPUT_LOAD,
    ZONES,
    FUELS,
    COMPONENT,
    REP_DAY_INT,
    LOAD_INC_CHG,
    TOTAL_LOAD_PLUS_EXPORTS,
    TOTAL_LOAD_PLUS_OVERGEN,
    DISPATCH_TIMESTAMP,
    COMPONENTTYPE,
    COMPONENT_NAME,
    COMPONENT_TYPE,
]


class ResolveHourlyResultsViewer:
    # Constants specific to ResolveResultsViewer
    ANNUAL = "annual"
    HOURLY = "hourly"
    CHRONO = "chrono"
    AGN_COL = "AnalysisGroupName"
    AGN_COLOR_COL = "AnalysisGroupColor"
    DEFAULT_AGG_GROUP_NAME = "edit analysis group name here"
    YEAR_HOUR = "year_hour"
    HOURS_PER_DAY = 24
    COMPONENT_TYPE_COLOR_COL = "COMPONENT_TYPE color"
    AGG_DF_KEY = "default_aggregation_df"
    DISPATCH_HOUR = "dispatch_hour"

    def __init__(
        self,
        case_results_folder: str,
        chrono_date_list: list[datetime.date] = None,
        modeled_years: list[int] = None,
        aggregation_config_df: pd.DataFrame = None,
        enable_column_filter_on_load: bool = True,
        additional_columns: list[str] = None,  # if additional columns aside from defaults are desired, use this arg
        enable_time_filter_on_load: bool = True,
        max_files: int = 10000,
        emit_dw_weights: Optional[bool] = False,
    ):
        """
        Initialize the ResolveResultsViewer.

        Args:
            case_results_folder (str): Path to the folder containing case results.
            chrono_date_list (list[datetime.date]): List of dates for filtering data.
            modeled_years (list[int], optional): List of vintage years. Defaults to None. If None, all modeled years available in data are used.
            aggregation_config_df (pd.DataFrame, optional): DataFrame containing aggregation configuration. Defaults to None.
            enable_column_filter_on_load (bool, optional): Whether to enable column filtering on load. Defaults to True.
            additional_columns (list[str], optional): List of additional columns to load. Defaults to None.
            enable_time_filter_on_load (bool, optional): Whether to enable time filtering on load. Defaults to True.
            max_files (int, optional): Maximum number of files to load. Defaults to 10000.
            emit_dw_weights (bool, optional): If True, include Dispatch Window "weight" col in outputs. Default False. "weight" is not used in aggregation, and is aligned to chrono timeline for convenience".

        """
        self.allowed_file_ingestion_columns = list(
            set(OP_COLS_SUPERSET + HOURLY_OP_COLS + CHRONO_OP_COLS + ADDITIONAL_COLUMNS)
        )
        self.case_results_folder = pathlib.Path(case_results_folder)
        self.temporal_settings_dict = self.get_temporal_settings_from_dir(case_results_folder)

        if chrono_date_list is None:
            all_dispatch_windows = [
                c.date() for c in set(self.temporal_settings_dict["chrono_periods_map"]["dispatch_window"])
            ]
            self.chrono_date_list = list(sorted(set(all_dispatch_windows)))
        else:
            self.chrono_date_list = list(sorted(set(chrono_date_list)))

        self.modeled_years = modeled_years or []
        self.enable_column_filter_on_load = enable_column_filter_on_load
        self.enable_time_filter_on_load = enable_time_filter_on_load
        self.additional_columns = additional_columns or []
        self.max_files = max_files
        self.emit_dw_weights = emit_dw_weights
        self.resolve_case_component_summary = self._load_component_summary()

        self.default_aggregation_df = self._get_default_aggregation_df()

        self.configured_aggregation_df = self._load_aggregation_config(aggregation_config_df)
        self.resolve_aggregation_df = self._reduce_aggregation_plan()

        if chrono_date_list is None:
            all_dispatch_windows = [
                c.date() for c in set(self.temporal_settings_dict["chrono_periods_map"]["dispatch_window"])
            ]
            self.chrono_date_list = list(sorted(set(all_dispatch_windows)))

        self.agg_color_map = self._get_agg_color_map_from_agg_plan(self.resolve_aggregation_df)

        self.chrono_timestamp_df = self.get_chrono_timestamp(self.chrono_date_list, self.temporal_settings_dict)
        self.component_summary_chronoextract_by_modelyear = self.get_component_summary_chronoextract_by_modelyear()
        available_requested_model_years = list(self.component_summary_chronoextract_by_modelyear[self.HOURLY].keys())
        if modeled_years is not None:
            years_requested_not_available = set(modeled_years) - set(available_requested_model_years)
            if years_requested_not_available:
                logger.warning(
                    f"Requested modeled years {years_requested_not_available} not available in {case_results_folder}\nProceeding with years {available_requested_model_years}"
                )
            else:
                logger.info(f"Using specified modeled years {available_requested_model_years}")
        else:
            logger.info(f"Using modeled years found in data {available_requested_model_years}")
        self.modeled_years = available_requested_model_years

        # TODO: Speed up the following dataframe joining or get rid of it if it's not necessary to merge all data
        # self.system_dispatch_by_resource = self.get_joined_chronoperiod_dispatch_per_modeledyears(self.modeled_years)

    ########################################
    ### CASE RESULTS LOADING AND JOINING ###
    ########################################

    def _load_component_summary(self) -> dict[str, pd.DataFrame]:
        resolve_case_res_grp_summary = {}
        for times_scheme in [self.ANNUAL, self.HOURLY, self.CHRONO]:
            try:
                resolve_case_res_grp_summary[times_scheme] = self._get_components_summary(times_scheme)
                if len(resolve_case_res_grp_summary[times_scheme]) == 0:
                    del resolve_case_res_grp_summary[times_scheme]  # Delete empty dataframes
            except ValueError as e:
                logger.warning(f"Could not load {times_scheme} data: {e}")
        return resolve_case_res_grp_summary

    def _get_components_summary(self, time_scheme: Literal["hourly", "chrono", "annual"]) -> pd.DataFrame:
        """Load CSV files from a Resolve case summary based on naming:
            *hourly_results.csv OR *chrono_results.csv OR *annual_results.csv

        Args:
            time_scheme (Literal["hourly", "chrono", "annual"]): Which CSV files to load

        Returns:
            pd.DataFrame: contains data from all matched CSV files that could be loaded
        """
        component_result_path = self.case_results_folder / "summary"

        if time_scheme == "annual":
            component_summary_result_files = glob.glob(
                f"**/*{time_scheme}_results_summary.csv", root_dir=pathlib.Path(component_result_path), recursive=True
            )
        else:
            component_summary_result_files = glob.glob(
                f"**/*{time_scheme}_results.csv", root_dir=pathlib.Path(component_result_path), recursive=True
            )
        if len(component_summary_result_files) > self.max_files:
            logger.error(
                f"Too many '{time_scheme}' files to load: {len(component_summary_result_files)}, max_files={self.max_files}\nTo proceed with this big job, re-run using a larger max_files parameter."
            )
        else:
            logger.info(
                f"reading {len(component_summary_result_files)} '{time_scheme}' files from '{component_result_path}')"
            )

        df_list = Parallel(n_jobs=-1)(
            delayed(self._read_and_process_file)(file, component_result_path)
            for file in tqdm(component_summary_result_files, desc="Loading component summary files")
        )

        cmp_sum_results = pd.concat(df_list, axis=0)
        if cmp_sum_results.empty:
            logger.warning(f"Empty DataFrame after loading all {time_scheme} files")
        else:
            logger.info(
                f"Loaded {time_scheme} data from {len(set(cmp_sum_results[COMPONENT]))} components, shape: {cmp_sum_results.shape}"
            )
        return cmp_sum_results

    def _read_and_process_file(self, file: str, component_result_path: pathlib.Path) -> pd.DataFrame:
        allowed_data_columns = self.allowed_file_ingestion_columns + self.additional_columns
        fp = component_result_path / file
        keep_any_col = not self.enable_column_filter_on_load
        df = pd.read_csv(fp, usecols=lambda x: (x.strip() in allowed_data_columns) or (keep_any_col))
        if not df.empty:
            df.rename(columns=lambda x: x.strip(), inplace=True)
            if "annual_results_summary" in file:
                df.rename(
                    columns={COMPONENT_TYPE: COMPONENTTYPE, COMPONENT_NAME: COMPONENT, MODELED_YEAR: MODELED_YEARS},
                    inplace=True,
                )
            else:
                df.loc[:, COMPONENTTYPE] = pathlib.Path(file).parts[-3]  # parent of deepest folder
                df.loc[:, COMPONENT] = pathlib.Path(file).parts[-2]  # deepest folder
            df1 = df.copy()

            # Fill in NaN values in certain model years for ZONES, FUELS
            if ZONES in df1.columns:
                component_zone_map = df1.dropna(subset=[ZONES]).drop_duplicates(COMPONENT).set_index(COMPONENT)[ZONES]
                df1[ZONES] = df1[COMPONENT].map(component_zone_map)
            if FUELS in df1.columns:
                component_fuel_map = df1.dropna(subset=[FUELS]).drop_duplicates(COMPONENT).set_index(COMPONENT)[FUELS]
                df1[FUELS] = df1[COMPONENT].map(component_fuel_map)

            if self.enable_time_filter_on_load:
                if len(self.modeled_years) > 0:
                    for year_header in [MODELEDYEAR, MODELED_YEARS]:
                        if year_header in df:
                            mask = pd.DatetimeIndex(df1[year_header]).year.isin(self.modeled_years)
                            df1 = df1[mask]
                for chrono_header in [CHRONO_PERIOD, CHRONO_PERIODS]:
                    if chrono_header in df:
                        idx = pd.DatetimeIndex(df1[chrono_header]).floor("D")
                        mask = idx.isin(self.chrono_date_list)
                        df1 = df1[mask]
                        break  # do not filter on both CHRONO_PERIOD and CHRONO_PERIODS
                if df1.empty:
                    logger.warning(f"Empty DataFrame after time filter for file:\n{file}")
                    result = pd.DataFrame()
                else:
                    result = df1
        else:
            cols = pd.read_csv(fp).columns
            logger.warning(f"Empty DataFrame after column filter for file:\n{file}\nwith available columns:\n{cols}")
            result = pd.DataFrame()
        return result

    def get_component_summary_chronoextract_by_modelyear(
        self,
    ) -> dict[str, dict[int, pd.DataFrame]]:
        """
        Returns a dict of modeled years

        "Data extracts to support plotting specified in
        Asana link: https://app.asana.com/0/1208669543793941/1208678988784415/f
        """
        # Build hourly timeline over the extract dates of chrono timeline
        chrono_timestamp_df = self.chrono_timestamp_df
        # get_chrono_timestamp(chrono_start_date, chrono_end_date, temporal_settings_dict)

        # Load Resolve run summary data across all Components
        res_case_cmp_sum = self.resolve_case_component_summary

        # Annual
        component_summary_chronoextract_by_modelyear = {}
        annual_df = res_case_cmp_sum[self.ANNUAL].copy()
        annual_df[MODELEDYEAR] = pd.DatetimeIndex(annual_df[MODELED_YEARS]).year.fillna(-1).astype(int)
        annual_df = self._strip_whitespace_df_string_values(annual_df, [COMPONENTTYPE, COMPONENT, ZONES])
        component_summary_chronoextract_by_modelyear[self.ANNUAL] = annual_df

        # Hourly
        if self.HOURLY in res_case_cmp_sum and len(res_case_cmp_sum[self.HOURLY]) > 0:
            hourly_extract = res_case_cmp_sum[self.HOURLY].copy()
            hourly_extract[TIMESTAMP] = pd.DatetimeIndex(hourly_extract[TIMESTAMPS])
            hourly_extract[MODELEDYEAR] = pd.DatetimeIndex(hourly_extract[MODELED_YEARS]).year.fillna(-1).astype(int)
            hourly_extract = self._strip_whitespace_df_string_values(hourly_extract, [COMPONENTTYPE, COMPONENT, ZONES])
            component_summary_chronoextract_by_modelyear[self.HOURLY] = dict(
                list(chrono_timestamp_df.reset_index().merge(hourly_extract, on=[TIMESTAMP]).groupby(MODELEDYEAR))
            )

        # Chrono
        if self.CHRONO in res_case_cmp_sum and len(res_case_cmp_sum[self.CHRONO]) > 0:
            chrono_extract = res_case_cmp_sum[self.CHRONO].copy()
            chrono_extract[TIMESTAMP] = pd.DatetimeIndex(chrono_extract[TIMESTAMPS])
            chrono_extract[CHRONO_PERIOD] = pd.DatetimeIndex(chrono_extract[CHRONO_PERIODS])
            chrono_extract[MODELEDYEAR] = pd.DatetimeIndex(chrono_extract[MODELED_YEARS]).year.fillna(-1).astype(int)
            chrono_extract = self._strip_whitespace_df_string_values(chrono_extract, [COMPONENTTYPE, COMPONENT, ZONES])
            component_summary_chronoextract_by_modelyear[self.CHRONO] = dict(
                list(
                    chrono_extract.merge(chrono_timestamp_df.reset_index(), on=[CHRONO_PERIOD, TIMESTAMP]).groupby(
                        MODELEDYEAR
                    )
                )
            )

        return component_summary_chronoextract_by_modelyear

    def get_hourly_results_for_all_years(self, modeled_years: list[int]) -> pd.DataFrame:
        """
        Combine the dictionary of hourly results by model year keys into a single dataframe, with DW weights
        Args:
            modeled_years: list of modeled years over which to concatenate

        Returns:
            Single dataframe of all hourly data
        """
        dispatch_window_weights = self.temporal_settings_dict[
            "dispatch_window_weights"
        ]  # TODO: Decide how weights are used in aggregation
        dww_df = dispatch_window_weights.set_index(DISPATCH_WINDOW)
        system_dispatch_by_resource = self.component_summary_chronoextract_by_modelyear
        if self.component_summary_chronoextract_by_modelyear.keys() >= {
            self.HOURLY,
        }:
            join_index = [
                c
                for c in [MODELEDYEAR, COMPONENTTYPE, COMPONENT, CHRONO_TIMESTAMP, DISPATCH_WINDOW, TIMESTAMP]
                if c in system_dispatch_by_resource[self.HOURLY][modeled_years[0]]
            ]

            df_h = pd.concat([system_dispatch_by_resource[self.HOURLY][year] for year in modeled_years]).set_index(
                join_index
            )

            df_h_with_weights = df_h.join(dww_df, on=DISPATCH_WINDOW)
            if df_h_with_weights[DISPATCH_WINDOW_WEIGHT].isna().any():
                ddf = (
                    df_h_with_weights.reset_index(DISPATCH_WINDOW)[[DISPATCH_WINDOW, DISPATCH_WINDOW_WEIGHT]]
                    .drop_duplicates()
                    .reset_index(drop=True)
                )
                unmatchedset = ddf.loc[ddf[DISPATCH_WINDOW_WEIGHT].isna()]
                logger.warning(
                    f"Missing weights for dispatch windows:\n{unmatchedset}\nProceeding with weight=0 for those dispatch windows"
                )
                df_h_with_weights[DISPATCH_WINDOW_WEIGHT] = df_h_with_weights[DISPATCH_WINDOW_WEIGHT].fillna(0.0)
            df_h = df_h_with_weights

            # Specify columns
            hourly_op_cols = HOURLY_OP_COLS + [DISPATCH_WINDOW_WEIGHT]  # Use the constant
            hourly_fields = join_index + hourly_op_cols
            hourly_fields_confirmed = df_h.reset_index().columns.intersection(hourly_fields)
            df_hourly = df_h.reset_index()[hourly_fields_confirmed].set_index(join_index).sort_index()

            return df_hourly
        else:
            raise ValueError("There are no hourly results reported in this case.")

    def get_joined_chronoperiod_dispatch_per_modeledyears(self, modeled_years: list[int]) -> pd.DataFrame:
        """
        Get the system dispatch data for specific days or weeks, grouped by resource type.
        Asana link: https://app.asana.com/0/1208669543793941/1208678988784415/f

        Args:
            modeled_years (list[int]): The years being modeled.

        Returns:
            pd.DataFrame: The DataFrame containing the dispatch data.
        """

        # TODO: Speed this up, it's very slow in large cases

        # HOURLY fields available
        # ['chrono_timestamp', 'dispatch_window', 'timestamp', 'include',
        # 'chrono_period', 'MODELED_YEARS', 'DISPATCH_WINDOWS', 'TIMESTAMPS',
        # 'weight', 'Zone(s)', 'Power Input (MW)', 'Power Output (MW)',
        # 'Net Power Output (MW)', 'Power Output Upper Bound (MW)',
        # 'Power Output Lower Bound (MW)', 'Total Up Reserves (MW)',
        # 'Total Down Reserves (MW)', 'Total Provide Reserve (MW)',
        # 'Power Output Max Dual', 'Power Output Variable Cost ($)',
        # 'Production Tax Credit ($)', 'Power Input Upper Bound (MW)',
        # 'Power Input Lower Bound (MW)', 'Power Input Variable Cost ($)',
        # 'Variable O&M Cost', 'Min Power Output Profile',
        # 'Max Power Output Profile', 'Variable O&M Cost.1',
        # 'Min Power Input Profile', 'Max Power Input Profile',
        # 'Charging Efficiency', 'Discharging Efficiency', 'ResourceGroup',
        # 'file_stem', 'file_path', 'modeled_year']

        # CHRONO fields available
        # ['chrono_timestamp', 'dispatch_window', 'timestamp', 'include',
        # 'chrono_period', 'MODELED_YEARS', 'CHRONO_PERIODS', 'TIMESTAMPS',
        # 'Zone(s)', 'SOC Inter-Intra Joint (MWh)', 'ResourceGroup', 'file_stem',
        # 'file_path', 'modeled_year']

        # Choose columns relevant for plotting
        system_dispatch_by_resource = self.component_summary_chronoextract_by_modelyear
        join_index = [
            c
            for c in [MODELEDYEAR, COMPONENTTYPE, COMPONENT, CHRONO_TIMESTAMP, DISPATCH_WINDOW, TIMESTAMP]
            if c in system_dispatch_by_resource[self.HOURLY][modeled_years[0]]
        ]
        chrono_op_cols = CHRONO_OP_COLS + [self.YEAR_HOUR]  # Use the constant
        chrono_fields = join_index + chrono_op_cols

        has_chrono = (
            self.CHRONO in system_dispatch_by_resource.keys() and len(system_dispatch_by_resource[self.CHRONO]) > 0
        )
        if has_chrono:
            df_c = pd.concat([system_dispatch_by_resource[self.CHRONO][year] for year in modeled_years]).set_index(
                join_index
            )
            df_chrono = df_c.reset_index()[chrono_fields].set_index(join_index).sort_index()
        else:
            df_chrono = pd.DataFrame()

        # Join chrono summary onto hourly summary
        logger.debug("Joining chrono summary onto hourly summary")
        df_hourly = self.get_hourly_results_for_all_years(modeled_years=modeled_years)
        extract_df = pd.concat([df_hourly, df_chrono], axis=1)

        # Get Zone and Fuel columns from Annual summary
        annual_fields = [MODELEDYEAR, ZONES, FUELS, OPERATIONAL_CAPACITY, OPERATIONAL_STORAGE_CAPACITY]
        df_annual = system_dispatch_by_resource[self.ANNUAL].set_index([COMPONENT])[annual_fields]
        final_index = extract_df.index
        extract_df = (
            extract_df.reset_index()
            .merge(df_annual, how="left", on=[MODELEDYEAR, COMPONENT])
            .set_index(final_index.names, drop=True)
        )

        # Reindex
        extract_df = extract_df.sort_index()
        new_idx = [c for c in [OPN_GROUP, FUELS, ZONES] if c in extract_df]
        extract_df = extract_df.set_index(new_idx, append=True)

        # Convenience for aggregation and plotting by YEAR_HOUR
        extract_df[self.YEAR_HOUR] = pd.DatetimeIndex(extract_df.reset_index()[CHRONO_TIMESTAMP]).strftime("%Y %Hh")

        return extract_df

    ##################################
    ### AGGREGATION CONFIG LOADING ###
    ##################################
    def _get_default_aggregation_df(self) -> pd.DataFrame:
        # HELPER METHODS
        def _prepare_aggregation_data(
            df: pd.DataFrame, default_agg_cols: list[str], op_cols_superset: list[str]
        ) -> Tuple[list[str], list[str], pd.DataFrame]:
            agg_cols = list(df.columns.intersection(default_agg_cols))
            op_cols = list(df.columns.intersection(op_cols_superset))
            df = df[agg_cols + op_cols].copy()
            df.loc[:, agg_cols] = df[agg_cols].fillna(-1)
            return agg_cols, op_cols, df

        def _create_default_agg_df(df: pd.DataFrame, agg_cols: list[str]) -> pd.DataFrame:
            df = df.set_index(agg_cols).sort_index().dropna(how="all", axis=0).reset_index()
            # Find minimal set of agg_cols rows that have at least one numeric value in op_cols
            idx_with_numeric_op_data = df.groupby(by=agg_cols).mean().abs().notna().any(axis=1)
            return idx_with_numeric_op_data.reset_index().iloc[:, 0:-1]

        def _create_aggregation_plan(result: dict[str, dict[str, pd.DataFrame]]) -> pd.DataFrame:
            """Matching strategy depends which of the three time schemes are available in the data."""
            AGG_DF_KEY = self.AGG_DF_KEY
            match_keys = set(result.keys())
            if match_keys == set([self.ANNUAL, self.HOURLY, self.CHRONO]):
                agg_df = (
                    pd.concat((result[self.HOURLY][AGG_DF_KEY], result[self.CHRONO][AGG_DF_KEY]), axis=0)
                    .fillna(-1)
                    .drop_duplicates()
                )
            elif match_keys == set([self.ANNUAL, self.HOURLY]):
                agg_df = result[self.HOURLY][AGG_DF_KEY].fillna(-1).drop_duplicates()
            elif match_keys == set([self.ANNUAL, self.CHRONO]):
                agg_df = result[self.CHRONO][AGG_DF_KEY].fillna(-1).drop_duplicates()
            elif match_keys == set([self.ANNUAL]):
                agg_df = result[self.ANNUAL][AGG_DF_KEY].fillna(-1).drop_duplicates()
            else:
                logger.warning(f"Could not find matching keys in {match_keys}")

            # Rows with operational group set correspond to components that are already counted in a group.
            # To avoid double counting, keep only rows with operational group not set (ie. -1)
            # After which, drop the operational group column because it is not a basis for grouping
            if OPN_GROUP in agg_df:
                default_aggregation_df = agg_df[agg_df[OPN_GROUP] == -1].drop(columns=OPN_GROUP)
            else:
                default_aggregation_df = agg_df

            # Drop MODELED_YEARS column because it is not a basis for aggregation (though vintage info may still exist in component names)
            if "MODELED_YEARS" in default_aggregation_df.columns:
                logger.info(f"Dropping {MODELED_YEARS} column from default aggregation plan")
                default_aggregation_df = default_aggregation_df.drop(columns="MODELED_YEARS").drop_duplicates()
            else:
                default_aggregation_df = default_aggregation_df.drop_duplicates()
            default_aggregation_df = default_aggregation_df.sort_values(
                by=list(default_aggregation_df.columns.drop(self.AGN_COL))
            )

            return default_aggregation_df

        # CREATE DEFAULT AGGREGATION DF
        df_dict = self.resolve_case_component_summary
        # Columns that might be considered as a basis for grouping
        default_agg_cols = [
            COMPONENTTYPE,
            COMPONENT,
            ZONES,
            FUELS,
        ]
        op_cols_superset = OP_COLS_SUPERSET  # Use the constant
        result = {}
        for k, v in df_dict.items():

            # Guard clause for empty dataframe
            if len(v) == 0:
                continue

            # Keep only agg_cols row-tuples with at least one numeric value in corresponding op_cols .
            # op_cols can be configured generously
            agg_cols, op_cols, df = _prepare_aggregation_data(v, default_agg_cols, op_cols_superset)
            default_agg_df = _create_default_agg_df(df, agg_cols)
            default_agg_df[self.AGN_COL] = self.DEFAULT_AGG_GROUP_NAME

            result |= {
                k: {
                    "agg_cols": agg_cols,
                    "op_cols": op_cols,
                    self.AGG_DF_KEY: default_agg_df,
                },
            }

        default_agg_df = _create_aggregation_plan(result)
        # Use component name as default aggregation group name.
        default_agg_df[self.AGN_COL] = default_agg_df[COMPONENT]
        return default_agg_df

    def _load_aggregation_config(self, aggregation_config_df: pd.DataFrame) -> pd.DataFrame:
        # TODO: Determine if this should be exported.
        # aggregation_settings_path = (
        #     self.case_results_folder.parent / "aggregation_settings" / self.case_results_folder.name
        # )
        # aggregation_settings_path.mkdir(exist_ok=True, parents=True)
        # fname = str(aggregation_settings_path / "resolve_aggregation_config.csv")
        # self.default_aggregation_df.to_csv(fname, index=False)
        # logger.info(
        #     f"Saved default aggregation plan to {fname}. (The default captures possible aggregations from data, as a starting point for user configuration.)"
        # )
        if aggregation_config_df is None:
            logger.info(f"Use default aggregation plan.")
            return self.default_aggregation_df
        else:
            # TODO: Throw a warning if there are resources in agg plan which are not in the case results, or vice versa
            logger.info("User provided aggregation plan dataframe.")
            return aggregation_config_df

    def _reduce_aggregation_plan(self) -> pd.DataFrame:
        available_agg_cols = [
            c
            for c in self.configured_aggregation_df.columns
            if c in self.default_aggregation_df.columns.drop(self.AGN_COL)
        ]
        if self.AGN_COLOR_COL in self.configured_aggregation_df.columns:
            resolve_aggregation_df = self.configured_aggregation_df[
                available_agg_cols + [self.AGN_COL, self.AGN_COLOR_COL]
            ].drop_duplicates()
        else:  # Color column is option to start
            resolve_aggregation_df = self.configured_aggregation_df[
                available_agg_cols + [self.AGN_COL]
            ].drop_duplicates()
        if set(resolve_aggregation_df.columns) != set(self.configured_aggregation_df.columns):
            logger.warning(
                f"Aggregation plan columns {set(resolve_aggregation_df.columns)} do not match configured columns {set(self.configured_aggregation_df.columns)}"
            )
        # TODO: Determine if it's necessary to write aggregation plans to csv
        # fname = str(
        #     pathlib.Path(self.case_results_folder.parent)
        #     / "aggregation_settings"
        #     / self.case_results_folder.name
        #     / "resolve_aggregation_actual.csv"
        # )
        # resolve_aggregation_df.to_csv(
        #     fname,
        #     index=False,
        # )
        # logger.info(f"Saved actual aggregation plan for this run to {fname}. Duplications are removed.")
        return resolve_aggregation_df

    #############################
    ### COLOR MAPPING METHODS ###
    #############################
    @staticmethod
    def _get_agg_color_map_from_agg_plan(agg_df: pd.DataFrame) -> dict[str, str]:
        """
        Generate a color map from the aggregation plan.

        Args:
            agg_df (pd.DataFrame): Aggregation DataFrame containing AGN_COL and AGN_COLOR_COL.

        Returns:
            dict[str, str]: A dictionary containing the color map.
        """
        if ResolveHourlyResultsViewer.AGN_COLOR_COL not in agg_df.columns:
            logger.warning(
                f"Generating arbitrary colors into missing column `{ResolveHourlyResultsViewer.AGN_COLOR_COL}` in aggregation plan"
            )
            agg_vals = agg_df[ResolveHourlyResultsViewer.AGN_COL].unique()
            num_colors = len(agg_vals)
            colors = px.colors.sample_colorscale(
                px.colors.qualitative.Plotly, [n / num_colors for n in range(num_colors)], colortype="tuple"
            )
            colors = [px.colors.convert_to_RGB_255(color) for color in colors]
            colors = [f"#{r:02x}{g:02x}{b:02x}" for r, g, b in colors]
            agg_color_map = {agg: color for agg, color in zip(agg_vals, colors)}
            agg_df[ResolveHourlyResultsViewer.AGN_COLOR_COL] = agg_df[ResolveHourlyResultsViewer.AGN_COL].map(
                agg_color_map
            )

        agg_color_map0 = agg_df[
            [ResolveHourlyResultsViewer.AGN_COL, ResolveHourlyResultsViewer.AGN_COLOR_COL]
        ].drop_duplicates()
        dups0 = agg_color_map0[agg_color_map0.duplicated(subset=[ResolveHourlyResultsViewer.AGN_COL], keep=False)]
        if not dups0.empty:
            logger.warning(
                f"Some {ResolveHourlyResultsViewer.AGN_COL} values have non-unique {ResolveHourlyResultsViewer.AGN_COLOR_COL} in this aggregation plan\n{dups0}\nProceeding with first color occurrence"
            )
        agg_color_map1 = agg_color_map0.drop_duplicates(subset=[ResolveHourlyResultsViewer.AGN_COL])
        dups1 = agg_color_map1[agg_color_map1.duplicated(subset=[ResolveHourlyResultsViewer.AGN_COLOR_COL], keep=False)]
        if not dups1.empty:
            logger.warning(
                f"Color mapping by {ResolveHourlyResultsViewer.AGN_COL} is not unique for this aggregation plan\n{dups1}\nProceeding non-unique colors; aggregation groups remain distinct in CSV files, and in plot keys, but not in plot colors"
            )
        aggregation_to_color_map = pd.Series(
            agg_color_map1[ResolveHourlyResultsViewer.AGN_COLOR_COL].values,
            index=agg_color_map1[ResolveHourlyResultsViewer.AGN_COL],
        ).to_dict()
        colors = set(aggregation_to_color_map.values())
        for color in colors:
            assert re.match(r"^#[0-9A-Fa-f]{6}$", color), f"Invalid color format: {color}"
        return aggregation_to_color_map

    def _ensure_unique_colors(self, map_df: pd.DataFrame) -> pd.DataFrame:
        if not map_df[ResolveHourlyResultsViewer.COMPONENT_TYPE_COLOR_COL].is_unique:
            map_df1 = map_df.sort_values(by=[ResolveHourlyResultsViewer.COMPONENT_TYPE_COLOR_COL])
            logger.warning("Color mapping by COMPONENT_TYPE only is not unique for this aggregation plan")
            delta_rgb = (
                np.min(
                    np.diff(
                        np.sort(
                            [int(x[1:], 16) for x in set(map_df1[ResolveHourlyResultsViewer.COMPONENT_TYPE_COLOR_COL])]
                        )
                    )
                )
                // 2
            )
            last_color = "#not_a_color"
            dup_count = 0
            for row in map_df1.itertuples():
                if row[2] == last_color:
                    last_color = row[2]
                    dup_count += 1
                    map_df1.at[row[0], ResolveHourlyResultsViewer.COMPONENT_TYPE_COLOR_COL] = (
                        f"#{int(row[2][1:],16)+delta_rgb*dup_count:x}".upper()
                    )
                else:
                    last_color = row[2]
                    dup_count = 0
            map_df = map_df1
        return map_df

    ##############################################
    ### RESOURCE AND ZONAL DISPATCH DATAFRAMES ###
    ##############################################
    def _aggregate_component_data(
        self,
        model_years: list[int],
        agg_time: str,
        df: pd.DataFrame,
        resolve_aggregation_df: pd.DataFrame,
        agg_fields: list[str],
        cmp_type_list: list[str],
        op_cols: list[str],
    ) -> pd.DataFrame:
        vvv = df[df[COMPONENTTYPE].isin(cmp_type_list)][
            [MODELEDYEAR, TIMESTAMP, agg_time] + agg_fields + [c for c in df.columns if c.strip() in op_cols]
        ]
        vvv = vvv.loc[vvv[MODELEDYEAR].isin(model_years)].dropna(how="all", axis=1)

        # Apply the AnalysisGroupName column
        uuu = vvv.merge(resolve_aggregation_df, on=agg_fields)
        if len(uuu) == 0:
            logger.warning(
                f"No data matches aggregation categories. Try collecting on these conditions: {vvv[resolve_aggregation_df.columns.drop(self.AGN_COL)].drop_duplicates()}"
            )

        # Perform the aggregation
        op_cols1 = [c for c in op_cols if c in uuu]
        missing_op_cols = set(op_cols) - set(op_cols1)
        if missing_op_cols:
            logger.warning(f"Missing operational columns {missing_op_cols} in {op_cols}")
        rrr = (
            uuu[[MODELEDYEAR, TIMESTAMP, agg_time, self.AGN_COL, *op_cols1]]
            .fillna(0)
            .groupby(by=[MODELEDYEAR, TIMESTAMP, agg_time, self.AGN_COL])
            .sum(numeric_only=True)[op_cols1]
            .reset_index()
        )
        return rrr

    def get_load_resource_balance(
        self,
        zone: str,
        model_years: list[int] = None,
        date_list: list[datetime.date] = None,
        agg_time: Literal["chrono_timestamp", "year_hour"] = "chrono_timestamp",
    ) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """
        Get the loads by component for specific model years and zone.

        Args:
            model_years (list[int]): The years being modeled.
            zone (str): The zone for which the data is retrieved.
            agg_time (Literal["chrono_timestamp", "year_hour"]): The aggregation time.

        Returns:
            Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]: The power provided DataFrame, zonal net imports DataFrame,
             and the load including charging DataFrame.
        """
        # If no modeled years are specified, LRB is generated for all years
        if model_years is None:
            model_years = self.modeled_years

        # TODO: year_hour aggregation must account for dispatch window weights or else averages are misleading
        if agg_time == self.YEAR_HOUR:
            raise NotImplementedError(f"{self.YEAR_HOUR} load-resource balance accounting is not implemented yet.")

        # Get dispatch of all resources
        df_hourly = self.get_hourly_results_for_all_years(modeled_years=model_years)
        hourly_index = df_hourly.index
        df_hourly = df_hourly.reset_index()
        annual_fields = [MODELEDYEAR, ZONES, FUELS, OPERATIONAL_CAPACITY, OPERATIONAL_STORAGE_CAPACITY]
        df_annual = self.component_summary_chronoextract_by_modelyear[self.ANNUAL].set_index([COMPONENT])[annual_fields]

        # Keep only rows of df that correspond to requested dates
        df_annual = df_annual.loc[df_annual[MODELEDYEAR].isin(model_years)]
        if date_list is not None:
            df_hourly = df_hourly.loc[df_hourly[CHRONO_TIMESTAMP].apply(lambda x: x.date()).isin(date_list)]

        # join hourly data to annual data
        df = (
            df_hourly.merge(df_annual, how="left", on=[MODELEDYEAR, COMPONENT])
            .set_index(hourly_index.names, drop=True)
            .reset_index()
        )

        # TODO: allow for a list of zones
        resolve_aggregation_df = self.resolve_aggregation_df.copy()

        # Keep only rows of resolve_aggregation_df that correspond to selected zone
        zonal_resources_df = resolve_aggregation_df.merge(df[[COMPONENT, ZONES]].drop_duplicates(), on=COMPONENT)
        zonal_resources_df = zonal_resources_df.loc[zonal_resources_df[ZONES] == zone]

        # Define fields on which to aggregate (typically ComponentType and Component)
        agg_fields = list(resolve_aggregation_df.columns.drop(self.AGN_COL).intersection(df.columns))

        # Prevent data mismatch due to whitespace
        df.loc[:, agg_fields] = df[agg_fields].fillna(-1).astype(str)  # -1 or "-1" is placeholder for NA in multi-index
        for col in agg_fields:
            df[col] = df[col].str.strip()

        # RESOURCE POWER PROVIDED
        cmp_type_list = [cls.__name__ for cls in GenericResource.get_subclasses()] + [GenericResource.__name__]
        op_cols = [POWER_OUT_MW, POWER_IN_MW, CURTAILED]
        power_provided_df = self._aggregate_component_data(
            model_years, agg_time, df, zonal_resources_df, agg_fields, cmp_type_list, op_cols
        )
        power_provided_df = power_provided_df.set_index([self.AGN_COL, agg_time]).sort_index().reset_index()

        # ZONAL NET IMPORTS
        cmp_type_list = ["Zone"]
        # Create aggregation plan for zonal net imports
        zonal_aggregation_df = pd.DataFrame([{"ComponentType": "Zone", "Component": zone, "AnalysisGroupName": zone}])
        op_cols = [
            POWER_OUT_MW,
            OVERGEN,
            UNSERVED_ENERGY,
            ZONAL_GROSS_IMPORTS,
            ZONAL_GROSS_EXPORTS,
            ZONAL_NET_IMPORTS,
            PRICES,
            CURTAILMENT,
        ]
        zonal_imports_df = self._aggregate_component_data(
            model_years, agg_time, df, zonal_aggregation_df, agg_fields, cmp_type_list, op_cols
        )

        # LOADS
        cmp_type_list = ["Zone"]
        op_cols = [
            POWER_IN_MW,
            INPUT_LOAD,
        ]
        load_summary_df = self._aggregate_component_data(
            model_years, agg_time, df, zonal_aggregation_df, agg_fields, cmp_type_list, op_cols
        )
        load_incl_chg_df = load_summary_df.set_index([self.AGN_COL, agg_time]).sort_index()
        load_incl_chg_df[ZONAL_TOTAL_LOAD] = load_incl_chg_df[POWER_IN_MW] + load_incl_chg_df[INPUT_LOAD]
        load_incl_chg_df = load_incl_chg_df.reset_index()

        # Append Dispatch window weights if requested.
        # Note: Over the WHOLE chrono timeline, the frequency of occurrence of rep days already happens in proportion
        # to the dispatch window weights, although that may not be evident in a short section of the chrono timeline.
        assert set(df.columns) >= {DISPATCH_WINDOW_WEIGHT}, f"Missing {DISPATCH_WINDOW_WEIGHT} column in {df.columns}"
        if self.emit_dw_weights:
            dww_df = df[[CHRONO_TIMESTAMP, DISPATCH_WINDOW_WEIGHT]].drop_duplicates().set_index(CHRONO_TIMESTAMP)
            power_provided_df1 = power_provided_df.join(dww_df, on=CHRONO_TIMESTAMP)
            assert (
                (power_provided_df == power_provided_df1.drop(columns=DISPATCH_WINDOW_WEIGHT)).all().all()
            ), "Dispatch window weights not joined correctly (power_provided_df)"
            power_provided_df = power_provided_df1
            zonal_imports_df1 = zonal_imports_df.join(dww_df, on=CHRONO_TIMESTAMP)
            assert (
                (zonal_imports_df == zonal_imports_df1.drop(columns=DISPATCH_WINDOW_WEIGHT)).all().all()
            ), "Dispatch window weights not joined correctly (zonal_imports_df)"
            zonal_imports_df = zonal_imports_df1
            load_incl_chg_df1 = load_incl_chg_df.join(dww_df, on=CHRONO_TIMESTAMP)
            assert (
                (load_incl_chg_df == load_incl_chg_df1.drop(columns=DISPATCH_WINDOW_WEIGHT)).all().all()
            ), "Dispatch window weights not joined correctly (load_incl_chg_df)"
            load_incl_chg_df = load_incl_chg_df1

        return power_provided_df, zonal_imports_df, load_incl_chg_df

    def _append_dispatch_window_weights_to_df(
        self, df_with_weights: pd.DataFrame, df_without_weights: pd.DataFrame
    ) -> pd.DataFrame:
        # Throw an error if dispatch window weight or chrono timestamp is not in the df_with_weights
        assert set(df_with_weights.columns) >= {
            DISPATCH_WINDOW_WEIGHT
        }, f"Missing {DISPATCH_WINDOW_WEIGHT} column in {df_with_weights.columns}"
        assert set(df_with_weights.columns) >= {
            CHRONO_TIMESTAMP
        }, f"Missing {CHRONO_TIMESTAMP} column in {df_with_weights.columns}"
        dww_df = (
            df_with_weights[[CHRONO_TIMESTAMP, DISPATCH_WINDOW_WEIGHT]].drop_duplicates().set_index(CHRONO_TIMESTAMP)
        )
        df_with_appended_weights = df_without_weights.join(dww_df, on=CHRONO_TIMESTAMP)
        assert (
            (df_without_weights == df_with_appended_weights.drop(columns=DISPATCH_WINDOW_WEIGHT)).all().all()
        ), "Dispatch window weights not joined correctly"
        return df_with_appended_weights

    def export_lrb_hourly_results(self, dest_path: str, modeled_years: list[int] = None, zones: [list[str]] = None):
        """Export the hourly zonal load and generation results to a specified destination path"""
        destination = pathlib.Path(f"{dest_path}/LRB Hourly Results")
        if not destination.exists():
            destination.mkdir(exist_ok=True)

        # Define zones for export
        if zones is None:
            zones_path = pathlib.Path(self.case_results_folder) / "summary" / "Zone"
            zones_for_export = [d.name for d in zones_path.iterdir() if d.is_dir()]
        else:
            zones_for_export = zones

        # Define modeled years for exporting hourly results
        if modeled_years is None:
            selected_modeled_years = self.modeled_years
        else:
            selected_modeled_years = modeled_years

        # Set emit DW weights equal to True
        self.emit_dw_weights = True

        power_df_list = []
        zonal_df_list = []
        logger.info(f"Getting load resource balance for modeled years: {selected_modeled_years}.")
        for zone in zones_for_export:
            logger.info(f"Getting load resource balance for zone {zone}.")
            # Get the LRB for all requested model years for this zone
            power_df, imports_df, load_df = self.get_load_resource_balance(
                zone=zone, model_years=selected_modeled_years, agg_time=CHRONO_TIMESTAMP
            )
            desired_index = ["Zone", "Modeled Year", "Dispatch Window Timestamp", "Dispatch Window Weight"]

            # clean up and write power output df
            power_df = power_df.rename(
                columns={
                    "AnalysisGroupName": "Resource",
                    "timestamp": "Dispatch Window Timestamp",
                    "modeled_year": "Modeled Year",
                    "weight": "Dispatch Window Weight",
                }
            )
            power_df["Zone"] = zone
            power_df = power_df.sort_values(by=["Modeled Year", "Dispatch Window Timestamp", "Resource"])
            power_df = power_df.set_index(desired_index)
            power_df = power_df.drop(columns=CHRONO_TIMESTAMP)
            power_df_list.append(power_df)

            # clean up and write zonal df
            zonal_df = load_df.merge(imports_df, how="left")
            zonal_df = zonal_df.drop(columns=CHRONO_TIMESTAMP)
            zonal_df = zonal_df.rename(
                columns={
                    "AnalysisGroupName": "Zone",
                    "timestamp": "Dispatch Window Timestamp",
                    "modeled_year": "Modeled Year",
                    "weight": "Dispatch Window Weight",
                }
            )
            zonal_df = zonal_df.set_index(desired_index)
            zonal_df_list.append(zonal_df)

        # Concat and save aggregated dataframes
        agg_power_df = pd.concat(power_df_list, ignore_index=False)
        agg_power_df.to_csv(destination / f"agg_hourly_power_output_by_resource.csv")
        agg_zonal_df = pd.concat(zonal_df_list, ignore_index=False)
        agg_zonal_df.to_csv(destination / f"agg_hourly_zonal_summary.csv")

        # Filter for zonal dataframes
        for zone in zones_for_export:
            zonal_path = destination / zone
            if not zonal_path.exists():
                zonal_path.mkdir(exist_ok=True)
            power_df = agg_power_df.loc[agg_power_df.index.get_level_values("Zone") == zone]
            zonal_df = agg_zonal_df.loc[agg_zonal_df.index.get_level_values("Zone") == zone]
            power_df.to_csv(zonal_path / f"{zone}_power_output_by_resource.csv")
            zonal_df.to_csv(zonal_path / f"{zone}_hourly_summary.csv")

    ##########################################
    ### AGGREGATED RESOURCE DISPATCH PLOTS ###
    ##########################################
    def _plot_line_agg_grp(self, agg_time: str, df: pd.DataFrame, y_col: str, name: str = None) -> Figure:
        """ "Include variable name in legend entries"""
        fig = px.line(df, x=agg_time, y=y_col, line_dash=self.AGN_COL)
        if name is not None:
            fig.update_traces(name=name)
        return fig

    def create_chrono_dispatch_plot(
        self,
        power_provided_df: pd.DataFrame,
        zonal_imports_df: pd.DataFrame,
        load_incl_chg_df: pd.DataFrame,
        plot_date_list: list[datetime.date] = None,
        ordered_names: list[str] = None,
        title_note: str = "Resource Dispatch Plot",
    ) -> Figure:
        """
        Create a dispatch plot for a specific chrono period.

        Args:
            power_provided_df (pd.DataFrame): The resource provide power df
            zonal_imports_df (pd.DataFrame): The zonal dataframe that includes imports, exports and unserved energy
            load_incl_chg_df (pd.DataFrame): The zonal dataframe that includes input load and total load
            plot_date_list (list[str]): The dates to plot
            ordered_names (list[str]): The order from bottom to top that the colors should appear on the chart
            title_note (str, optional): Additional note for the plot title. Defaults to "".

        Returns:
            Tuple[Figure, pd.DataFrame, pd.DataFrame]: The plot figure, power provided DataFrame, and load including charging DataFrame.
        """
        # All plots are in chrono time
        agg_time = CHRONO_TIMESTAMP
        # Filter power_provided_df, zonal_imports_df, and load_df based on specified list of days for plotting
        if plot_date_list is not None:
            power_provided_df = power_provided_df.loc[
                power_provided_df[CHRONO_TIMESTAMP].apply(lambda x: x.date()).isin(plot_date_list)
            ]
            zonal_imports_df = zonal_imports_df.loc[
                zonal_imports_df[CHRONO_TIMESTAMP].apply(lambda x: x.date()).isin(plot_date_list)
            ]
            load_incl_chg_df = load_incl_chg_df.loc[
                load_incl_chg_df[CHRONO_TIMESTAMP].apply(lambda x: x.date()).isin(plot_date_list)
            ]

        # Append curtailment, gross imports, and unserved energy to power_provided_df
        zonal_power_output = zonal_imports_df[POWER_OUT_MW]
        zonal_imports_df = zonal_imports_df.drop(columns=POWER_OUT_MW)

        # Imports
        zonal_imports_df[IMPORTS] = zonal_imports_df[ZONAL_NET_IMPORTS].clip(lower=0)
        zonal_imports_renamed = (
            zonal_imports_df.copy()
            .rename(columns={IMPORTS: POWER_OUT_MW})
            .assign(
                **{self.AGN_COL: IMPORTS},
            )
        )
        zonal_imports_renamed = zonal_imports_renamed[[agg_time, self.AGN_COL, POWER_OUT_MW]]

        # Unserved Energy
        zonal_unserved_energy = (
            zonal_imports_df.copy()
            .rename(columns={UNSERVED_ENERGY: POWER_OUT_MW})
            .assign(
                **{self.AGN_COL: "Unserved Energy"},
            )
        )
        zonal_unserved_energy = zonal_unserved_energy[[agg_time, self.AGN_COL, POWER_OUT_MW]]

        # Curtailment
        zonal_curtailment = (
            zonal_imports_df.copy()
            .rename(columns={CURTAILMENT: POWER_OUT_MW})
            .assign(
                **{self.AGN_COL: "Curtailment"},
            )
        )
        zonal_curtailment = zonal_curtailment[[agg_time, self.AGN_COL, POWER_OUT_MW]]

        power_provided_w_additions = power_provided_df.copy()
        if zonal_imports_renamed[POWER_OUT_MW].sum() > 0:
            power_provided_w_additions = pd.concat([power_provided_w_additions, zonal_imports_renamed])
        if zonal_unserved_energy[POWER_OUT_MW].sum() > 0:
            power_provided_w_additions = pd.concat([power_provided_w_additions, zonal_unserved_energy])
        power_provided_w_additions = pd.concat([power_provided_w_additions, zonal_curtailment])

        # Hard-coded colors for curtailment, imports and unserved energy
        agg_color_map_w_additions = self.agg_color_map.copy()
        agg_color_map_w_additions["Curtailment"] = "#DDD9C3"
        agg_color_map_w_additions["Imports"] = "#FFC9BB"
        agg_color_map_w_additions["Unserved Energy"] = "#FF0000"  # red for unserved energy

        # Stacked area plot for power provided (including zonal net imports)
        kwargs = dict(
            x=agg_time,
            y=POWER_OUT_MW,
            color=self.AGN_COL,
            color_discrete_map=agg_color_map_w_additions,
        )
        # Add AGN_COL order if specified
        if ordered_names is not None:
            kwargs["category_orders"] = {self.AGN_COL: ordered_names}
        fig = px.area(power_provided_w_additions, **kwargs)
        # Remove borders (line width) on each area trace
        for trace in fig.data:
            trace.update(line=dict(width=0))
            # Set full opacity for each trace to remove the pastel look
            base_color = trace.line.color  # The original hex color assigned by px.area
            trace.update(
                fillcolor=base_color,  # force full hex color with no transparency
                line=dict(width=0),  # optional: remove border if you want a cleaner look
            )

        # Plot both input load and total load as a line of different colors
        fig2 = self._plot_line_agg_grp(agg_time, load_incl_chg_df, INPUT_LOAD, name="Input Load")
        fig2.update_traces(showlegend=True, line_color="blue", patch={"line": {"dash": "dash"}})
        fig3 = self._plot_line_agg_grp(
            agg_time, load_incl_chg_df, ZONAL_TOTAL_LOAD, name="Input Load + Storage Charging Load"
        )
        fig3.update_traces(
            showlegend=True,
            line_color="red",
        )
        # Plot gross exports plus total load as a third line (only if gross exports are non-zero)
        load_incl_chg_and_exports = load_incl_chg_df.copy()
        load_incl_chg_and_exports[TOTAL_LOAD_PLUS_EXPORTS] = load_incl_chg_and_exports[ZONAL_TOTAL_LOAD] + (
            -1 * zonal_imports_df[ZONAL_NET_IMPORTS].clip(upper=0)
        )
        if (
            load_incl_chg_and_exports[TOTAL_LOAD_PLUS_EXPORTS].sum()
            != load_incl_chg_and_exports[ZONAL_TOTAL_LOAD].sum()
        ):
            fig4 = self._plot_line_agg_grp(
                agg_time,
                load_incl_chg_and_exports,
                TOTAL_LOAD_PLUS_EXPORTS,
                name="Input Load + Charging Load + Exports",
            )
            fig4.update_traces(
                showlegend=True,
                line_color="black",
            )
        # Check if there is any overgeneration (very rare). If so, create another line that includes overgen
        if zonal_imports_df[OVERGEN].sum() > 0:
            load_incl_overgen = load_incl_chg_df.copy()
            load_incl_overgen[TOTAL_LOAD_PLUS_OVERGEN] = (
                load_incl_overgen[ZONAL_TOTAL_LOAD] + zonal_imports_df[ZONAL_GROSS_EXPORTS] + zonal_imports_df[OVERGEN]
            )
            fig5 = self._plot_line_agg_grp(
                agg_time, load_incl_overgen, TOTAL_LOAD_PLUS_OVERGEN, name="Input Load + Charging + Exports + Overgen"
            )
            fig5.update_traces(showlegend=True, line_color="black", patch={"line": {"dash": "dash"}})

        # Combine plots with correct draw order
        subfig = make_subplots()
        line_traces = []

        # Add stacked area traces first (drawn below, legend at bottom)
        for trace in fig.data:
            subfig.add_trace(trace)

        # Add fig5 and fig4 trace(s) if they exist
        if "fig5" in locals():
            for trace in fig5.data:
                trace.showlegend = False
                subfig.add_trace(trace)
                line_traces.append(("fig5", trace))
        if "fig4" in locals():
            for trace in fig4.data:
                trace.showlegend = False
                subfig.add_trace(trace)
                line_traces.append(("fig4", trace))
        for trace in fig3.data:
            trace.showlegend = False
            subfig.add_trace(trace)
            line_traces.append(("fig3", trace))
        for trace in fig2.data:
            trace.showlegend = False
            subfig.add_trace(trace)
            line_traces.append(("fig2", trace))

        # Add dummy legend traces in reverse draw order
        dummy_legend_order = ["fig2", "fig3", "fig4", "fig5"]  # reverse of draw order
        if "fig5" not in locals():
            dummy_legend_order.remove("fig5")
        if "fig4" not in locals():
            dummy_legend_order.remove("fig4")

        for name in dummy_legend_order:
            for trace_name, real_trace in line_traces:
                if trace_name == name:
                    dummy = go.Scatter(
                        x=[None],
                        y=[None],
                        mode=real_trace.mode,
                        name=real_trace.name,
                        line=real_trace.line,
                        showlegend=True,
                    )
                    subfig.add_trace(dummy)

        # Update layout
        subfig.update_layout(
            width=1200,
            height=600,
            font=dict(family="Arial", size=18, color="black"),
            title={"text": title_note, "font": {"size": 20, "family": "Arial Black"}},
            xaxis_title="Hour Beginning",
            xaxis=dict(
                dtick=3600000,
                tickformat="%H",
                showgrid=False,
                zeroline=False,
                ticklabelposition="outside bottom",
                tickangle=0,
            ),
            yaxis_title={"text": "MW"},
            yaxis=dict(
                tickformat=",",
                showgrid=False,
                zeroline=False,
            ),
            plot_bgcolor="white",
            paper_bgcolor="white",
            margin=dict(b=100),
            legend_traceorder="reversed",
        )

        return subfig

    ######################################
    ### STATE OF CHARGE TRACKING PLOTS ###
    ######################################
    def determine_soc_source(self, df: pd.DataFrame) -> Optional[str]:
        """
        Determine which State of Charge (SOC) column to use based on data availability.

        Args:
            df (pd.DataFrame): DataFrame containing potential SOC columns

        Returns:
            Optional[str]: Name of the SOC column to use, or None if no valid SOC data found
                Returns `SOC Inter-Intra Joint (MWh)` if both columns have data
                Returns `SOC Intra Period (MWh)` if only intra-period data exists
                Returns `SOC Inter-Intra Joint (MWh)` if only inter-intra joint data exists
                Returns None if no SOC data exists
        """
        # Check for SOC_INTRA_MWh data
        has_soc_intra_mwh = SOC_INTRA_MWh in df.columns and df[SOC_INTRA_MWh].ne(0).any()

        # Check for SOC_IIJ_MWh data
        has_soc_iij_mwh = SOC_IIJ_MWh in df.columns and df[SOC_IIJ_MWh].ne(0).any()

        # Determine which column to use
        if has_soc_intra_mwh and has_soc_iij_mwh:
            return SOC_IIJ_MWh
        elif has_soc_intra_mwh:
            return SOC_INTRA_MWh
        elif has_soc_iij_mwh:
            return SOC_IIJ_MWh
        return None

    def get_resource_soc_dataframe(
        self, resource_name: str, model_years: list[int] = None, date_list: list[datetime.date] = None
    ) -> pd.DataFrame:
        """
        Create a dataframe for resource or aggregated resource state of charge over time

        Args:
            resource_name: Name of resource or aggregated resource AnalysisGroupName
            model_years: The model years of the SOC to return (defaults to all of them)
            date_list: dates in chronological time for which to return the SOC (defaults to all of them)

        Returns:
            pd.DataFrame: chrono_timestamp indexed dataframe with resource or resource group SOC
        """
        # TODO: Finish the implementation of SOC by resource using this Aggregation Class.
        raise NotImplementedError(f"SOC by resource functionality not yet implemented.")

        # If no modeled years are specified, SOC df is generated for all years
        if model_years is None:
            model_years = self.modeled_years

        # Get dispatch of all resources
        df_hourly = self.get_hourly_results_for_all_years(modeled_years=model_years)
        hourly_index = df_hourly.index
        df_hourly = df_hourly.reset_index()
        annual_fields = [MODELEDYEAR, ZONES, FUELS, OPERATIONAL_CAPACITY, OPERATIONAL_STORAGE_CAPACITY]
        df_annual = self.component_summary_chronoextract_by_modelyear[self.ANNUAL].set_index([COMPONENT])[annual_fields]

        # Keep only rows of df that correspond to requested dates
        df_annual = df_annual.loc[df_annual[MODELEDYEAR].isin(model_years)]
        if date_list is not None:
            df_hourly = df_hourly.loc[df_hourly[CHRONO_TIMESTAMP].apply(lambda x: x.date()).isin(date_list)]

        # Get chrono results of all resources
        system_dispatch_by_resource = self.component_summary_chronoextract_by_modelyear
        chrono_op_cols = CHRONO_OP_COLS
        chrono_fields = hourly_index.names + chrono_op_cols
        has_chrono = (
            self.CHRONO in system_dispatch_by_resource.keys() and len(system_dispatch_by_resource[self.CHRONO]) > 0
        )
        if has_chrono:
            df_c = pd.concat([system_dispatch_by_resource[self.CHRONO][year] for year in model_years]).set_index(
                hourly_index
            )
            df_chrono = df_c.reset_index()[chrono_fields].set_index(hourly_index).sort_index()
        else:
            df_chrono = pd.DataFrame()

        # Get aggregation plan
        resolve_aggregation_df = self.resolve_aggregation_df.copy()

        # Filter df_chrono and df_hourly by the requested resource, join them
        if (resolve_aggregation_df[self.AGN_COL].eq(resource_name)).any():  # resource_name is an aggregation
            resolve_aggregation_df = resolve_aggregation_df.loc[resolve_aggregation_df[self.AGN_COL] == resource_name]
            resources_belonging_to_aggregated_storage = resolve_aggregation_df[COMPONENT].tolist()
            df_hourly = df_hourly.loc[df_hourly[COMPONENT].isin(resources_belonging_to_aggregated_storage)]
            if not df_chrono.empty:
                df_chrono = df_chrono.loc[df_chrono[COMPONENT].isin(resources_belonging_to_aggregated_storage)]
                df = pd.concat([df_hourly, df_chrono])
            else:
                df = df_hourly
        elif (resolve_aggregation_df[COMPONENT].eq(resource_name)).any():  # resource_name is a particular component
            resolve_aggregation_df = resolve_aggregation_df.loc[resolve_aggregation_df[COMPONENT] == resource_name]
            df_hourly = df_hourly.loc[df_hourly[COMPONENT] == resource_name]
            if not df_chrono.empty:
                df_chrono = df_chrono.loc[df_chrono[COMPONENT] == resource_name]
                df = pd.concat([df_hourly, df_chrono])
            else:
                df = df_hourly
        else:
            raise KeyError(f"Resource {resource_name} does not exist as either a component or a resource aggregation.")

        # join the aggregation df and the joined chrono and hourly df
        resource_df = resolve_aggregation_df.merge(df[[COMPONENT]].drop_duplicates(), on=COMPONENT)

        # Define fields on which to aggregate (typically ComponentType and Component)
        agg_fields = list(resolve_aggregation_df.columns.drop(self.AGN_COL).intersection(df.columns))

        # Prevent data mismatch due to whitespace
        df.loc[:, agg_fields] = df[agg_fields].fillna(-1).astype(str)  # -1 or "-1" is placeholder for NA in multi-index
        for col in agg_fields:
            df[col] = df[col].str.strip()
            logger.debug(f"{col} takes values : {set(df[col])}")

        # Get correct SOC column
        soc_col = self.determine_soc_source(df)
        if soc_col is None:
            raise ValueError(f"This resource doesn't have any SOC results for the selected timeframe.")

        # Create SOC df
        op_cols = [soc_col, POWER_OUT_MW, POWER_IN_MW]
        cmp_type_list = [cls.__name__ for cls in GenericResource.get_subclasses()] + [GenericResource.__name__]
        resource_soc_df = self._aggregate_component_data(
            model_years, CHRONO_TIMESTAMP, df, resource_df, agg_fields, cmp_type_list, op_cols
        )

        # Append Dispatch window weights if requested.
        # Note: Over the WHOLE chrono timeline, the frequency of occurrence of rep days already happens in proportion
        # to the dispatch window weights, although that may not be evident in a short section of the chrono timeline.
        if self.emit_dw_weights:
            resource_soc_df = self._append_dispatch_window_weights_to_df(
                df_with_weights=df, df_without_weights=resource_soc_df
            )

        return resource_soc_df

    def create_chrono_soc_plot(
        self,
        soc_df: pd.DataFrame,
        title_note: str = "",
    ) -> Tuple[Figure, pd.DataFrame, pd.DataFrame]:
        """
        Create a plot for State of Charge (SoC) over chrono time.

        Args:
            model_year (int): The year being modeled.
            zone (str): The zone for which the plot is created.
            title_note (str, optional): Additional note for the plot title. Defaults to "".

        Returns:
            Tuple[Figure, pd.DataFrame, pd.DataFrame]: The plot figure, power provided DataFrame, and load including charging DataFrame.
        """
        soc_col = self.determine_soc_source(soc_df)
        fig = self._plot_line_agg_grp(CHRONO_TIMESTAMP, soc_df, soc_col)
        fig.update_layout(title=title_note)
        return fig

    #############################
    ### HELPER STATIC METHODS ###
    #############################
    @staticmethod
    def _strip_whitespace_df_string_values(df: pd.DataFrame, string_data_cols: Iterable[str]) -> pd.DataFrame:
        strip_fail_cols = []
        for c in string_data_cols:
            try:
                df[c] = df[c].str.strip()
            except (AttributeError, KeyError):
                strip_fail_cols += [c]
        return df

    @staticmethod
    def get_chrono_timestamp(date_list: list[datetime.date], temporal_settings_dict: dict) -> pd.DataFrame:
        """Builds granular chrono timeline over selected chrono date interval from rep date mapping and timestamps"""
        u = temporal_settings_dict["chrono_periods_map"].set_index(CHRONO_PERIOD).loc[date_list]
        m = temporal_settings_dict["dispatch_windows_map"].set_index(DISPATCH_WINDOW)
        w = (
            m.reset_index()
            .merge(u.reset_index(), on=DISPATCH_WINDOW, how="inner")
            .sort_values(by=[CHRONO_PERIOD, TIMESTAMP])
            .set_index(CHRONO_PERIOD, drop=False)
        )
        w[CHRONO_TIMESTAMP] = w.index + (w.timestamp - pd.DatetimeIndex(w.timestamp).floor("D"))

        # Representative day, as an integer
        idx, _ = pd.factorize(w[DISPATCH_WINDOW])
        w[REP_DAY_INT] = idx

        # Convenience for aggregation and plotting by YEAR_HOUR
        w[ResolveHourlyResultsViewer.YEAR_HOUR] = pd.DatetimeIndex(w[CHRONO_TIMESTAMP]).strftime("%Y %Hh")

        chrono_timestamp_df = w.set_index(CHRONO_TIMESTAMP)
        return chrono_timestamp_df

    @staticmethod
    def get_temporal_settings_from_dir(case_results_folder) -> dict[str, pd.DataFrame]:
        """
        Returns temporal settings as a dict of dataframes loaded from filesystem CSVs.
        """

        temporal_settings_path = pathlib.Path(case_results_folder) / "temporal_settings"
        temporal_settings_dict = {}

        # Parse this files in the temporal_settings folder of the Resolve case output (as opposed to the Resolve case project folder)

        temporal_settings_dict["chrono_periods_map"] = pd.read_csv(
            temporal_settings_path / "chrono_periods_map.csv",
            parse_dates=[CHRONO_PERIOD, DISPATCH_WINDOW],
            dayfirst=False,
        )

        temporal_settings_dict["dispatch_window_weights"] = pd.read_csv(
            temporal_settings_path / "dispatch_window_weights.csv",
            parse_dates=[DISPATCH_WINDOW],
            dayfirst=False,
        )
        assert {DISPATCH_WINDOW, DISPATCH_WINDOW_WEIGHT}.issubset(
            temporal_settings_dict["dispatch_window_weights"].columns
        )

        temporal_settings_dict["modeled_years"] = pd.read_csv(
            temporal_settings_path / "modeled_years.csv",
            parse_dates=[TIMESTAMP],
            dayfirst=False,
        )
        assert {"timestamp", "value"}.issubset(temporal_settings_dict["modeled_years"].columns)

        temporal_settings_dict["dispatch_windows_map"] = pd.read_csv(
            temporal_settings_path / "dispatch_windows_map.csv",
            parse_dates=[DISPATCH_WINDOW, TIMESTAMP],
            dayfirst=False,
        )
        assert {DISPATCH_WINDOW, TIMESTAMP, "include"}.issubset(temporal_settings_dict["dispatch_windows_map"].columns)

        temporal_settings_dict["modeled_year_discount_factors"] = pd.read_csv(
            temporal_settings_path / "modeled_year_discount_factors.csv"
        )

        # Some ingested files had trailing whitespace in column names
        for key in temporal_settings_dict:
            temporal_settings_dict[key].rename(columns=lambda x: x.strip(), inplace=True)

        dispatch_window_count = temporal_settings_dict["dispatch_window_weights"].shape[0]
        modeled_years_count = temporal_settings_dict["modeled_years"].value.value_counts()[True]
        logger.info(
            f"Temporal settings loaded: {modeled_years_count=}, {dispatch_window_count=}, {temporal_settings_dict.keys()=}"
        )

        return temporal_settings_dict

    @staticmethod
    def get_range_data_from_excel_rv(rv_path: str, named_range: str) -> pd.DataFrame:

        wb = openpyxl.load_workbook(rv_path, data_only=True, read_only=True)
        if named_range not in wb.defined_names:
            if named_range == HOURLY_AGG_SETTINGS_NAMED_RANGE:
                raise KeyError(
                    f"The named range {HOURLY_AGG_SETTINGS_NAMED_RANGE} does not exist in the Excel workbook at "
                    f"{rv_path}. You must define this named range to import the aggregation settings."
                )
            elif named_range == COLOR_SETTINGS_NAMED_RANGE:
                logger.warning(
                    f"The named range {COLOR_SETTINGS_NAMED_RANGE} does not exist in the Excel workbook at "
                    f"{rv_path}. If you plan to produce dispatch plots, you should define colors for each "
                    f"group within the Excel workbook with this named range."
                )
                return None  # Allow user to proceed without color settings in case charts are not needed
        dn = wb.defined_names[named_range]
        destinations = list(dn.destinations)

        # Make sure it's a single sheet
        sheets = {sheet for sheet, _ in destinations}
        if len(sheets) != 1:
            raise ValueError("Named range must be on one sheet.")
        sheetname = sheets.pop()
        sheet = wb[sheetname]

        columns = {}
        headers = {}

        # If single block -> one destination
        if len(destinations) == 1:
            _, cell_range = destinations[0]
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)

            # Read block once
            block = sheet.iter_rows(
                min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True
            )

            block = list(block)
            headers_row = block[0]
            data_rows = block[1:]

            for idx, header in enumerate(headers_row):
                columns[idx] = [row[idx] for row in data_rows]
                headers[idx] = header
        else:
            # Multiple separate destinations
            needed_cols = {}
            for _, cell_range in destinations:
                min_col, min_row, max_col, max_row = range_boundaries(cell_range)
                needed_cols[min_col] = (min_row, max_row)

            data_rows = list(sheet.iter_rows(values_only=True))

            for col_idx, (min_row, max_row) in needed_cols.items():
                col_data = [data_rows[row_idx - 1][col_idx - 1] for row_idx in range(min_row, max_row + 1)]
                headers[col_idx] = col_data[0]
                columns[col_idx] = col_data[1:]

        # Normalize lengths
        max_length = max(len(col) for col in columns.values())
        for col in columns.values():
            col.extend([None] * (max_length - len(col)))

        # Build DataFrame
        sorted_cols = sorted(columns.keys())
        df_data = {headers[col_idx]: columns[col_idx] for col_idx in sorted_cols}
        df = pd.DataFrame(df_data).dropna()

        # Normalize headers
        if named_range == HOURLY_AGG_SETTINGS_NAMED_RANGE:
            colnames = df.columns.tolist()
            df = df.rename(
                columns={colnames[0]: COMPONENT, colnames[1]: COMPONENTTYPE, colnames[2]: "AnalysisGroupName"}
            )
        elif named_range == COLOR_SETTINGS_NAMED_RANGE:
            colnames = df.columns.tolist()
            df = df.rename(
                columns={colnames[0]: "AnalysisGroupName", colnames[1]: "ChartOrder", colnames[2]: "AnalysisGroupColor"}
            )
            df = df.sort_values(by="ChartOrder")

        return df

    @staticmethod
    def combine_rv_groupings_and_color_settings(
        rv_groupings: pd.DataFrame, color_settings_df: pd.DataFrame
    ) -> pd.DataFrame:
        try:
            df = pd.merge(rv_groupings, color_settings_df, on=color_settings_df.columns[0])
        except KeyError:
            raise KeyError(
                "Specified grouping column in color settigns does not match specified grouping for aggregation."
            )
        df = df.drop(columns="ChartOrder")

        # Check if ComponentType only contains proper component types
        # Give a warning if any rows are removed
        all_component_subclasses = [cls.__name__ for cls in Component.get_subclasses()] + [Component.__name__]
        component_types_removed = df[~df[COMPONENTTYPE].isin(all_component_subclasses)][COMPONENTTYPE].unique().tolist()
        if component_types_removed:
            logger.warning(
                f"The following ComponentTypes were removed because they were not recognized: {component_types_removed}"
            )

        # filter out those rows
        df = df[df[COMPONENTTYPE].isin(all_component_subclasses)]

        return df

    ################################
    ### REP DAY STATISTICS PLOTS ###
    ################################
    def create_rep_day_hourly_plot(
        self, dispatch_df: pd.DataFrame, soc_col: str, title_msg: str
    ) -> Tuple[pd.DataFrame, Figure]:
        """
        Create a plot for cluster day, simulated day, and hourly State of Charge (SoC) tracking.
        Asana link: https://app.asana.com/0/1208669543793941/1208678988784415/f

        Args:
            title_msg (str): Title message for the plot.

        Returns:
            Tuple[pd.DataFrame, Figure]: The DataFrame for representative day hourly data and the plot figure.
        """

        title_msg += f"<br>{soc_col}"

        # plot generation for hourly SoC tracking
        df2 = dispatch_df[soc_col].reset_index()
        fig = px.line(
            df2,
            x=self.DISPATCH_HOUR,
            y=df2.columns,
            title=title_msg,
        )

        return fig

    def extract_dispatches_by_rep_day(self, dispatch_df: pd.DataFrame) -> Tuple[str, pd.DataFrame]:
        soc_col = self.determine_soc_source(dispatch_df)

        df = self.apply_dispatch_hourly_timestamps(dispatch_df, soc_col)

        df1 = df[[soc_col]].groupby([self.DISPATCH_HOUR]).describe()
        return soc_col, df1

    def apply_dispatch_hourly_timestamps(self, dispatch_df: pd.DataFrame, col: str) -> pd.DataFrame:
        """Infers columns `dispatch_timestamp` and `dispatch_hour` to DataFrame based on the date `dispatch_window`.
        Helpful for granular plotting and representative day analysis.
        """
        df = dispatch_df[
            [
                CHRONO_PERIOD,
                col,
            ]
        ].reset_index(DISPATCH_WINDOW)

        df = df.dropna(how="any", axis=0)  # TODO: manage DISPATCH_WINDOW = NaT in system_dispatch_by_resource
        df = df.reset_index(CHRONO_TIMESTAMP)
        df[DISPATCH_TIMESTAMP] = df[DISPATCH_WINDOW] + (df[CHRONO_TIMESTAMP] - df[CHRONO_PERIOD])

        df[self.DISPATCH_HOUR] = pd.DatetimeIndex(df[DISPATCH_TIMESTAMP]).strftime("%Y-%m-%d %Hh")
        # df.reset_index()[DISPATCH_TIMESTAMP].strftime()
        df = df.set_index(self.DISPATCH_HOUR, append=True)
        return df

    ###############################
    ### NOT IMPLEMENTED METHODS ###
    ###############################

    def create_yearhour_dispatch_plot(
        self,
        power_provided_df: pd.DataFrame,
        load_incl_chg_df: pd.DataFrame,
        yearhour_dispatch_per_resource_df: pd.DataFrame,
        modeled_year: int,
        zone: str,
        title_note: str,
    ) -> Figure:
        """
        Create an annual average dispatch plot that summarizes the annual behavior of resource types.

        Args:
            modeled_year (int): The year being modeled.
            zone (str): The zone for which the plot is created.
            title_note (str): Additional note for the plot title.

        Returns:
            Figure: Plot of dispatch aggregated to year-hour.
        """
        # """# Annual average dispatch charts that summarize the annual behaviour of resource types
        # Asana link: https://app.asana.com/0/1208669543793941/1208678988784415/f
        # """

        raise NotImplementedError(f"Year-hour plots are not yet implemented.")

        # agg_time = self.YEAR_HOUR
        # # power_provided_df, load_incl_chg_df, yearhour_dispatch_per_resource_df = self.get_yearhour_dispatch(modeled_year, zone)
        #
        # # plot generation
        # fig = px.bar(
        #     yearhour_dispatch_per_resource_df.reset_index(),
        #     x=agg_time,
        #     y=POWER_OUT_MW,
        #     color=self.AGN_COL,
        #     color_discrete_map=self.agg_color_map,
        #     title="Plot 2: " + title_note,
        #     barmode="stack",
        # )
        #
        # fig2 = self._plot_line_agg_grp(agg_time, power_provided_df, POWER_OUT_MW)
        # fig2.update_traces(showlegend=True, line_color="blue")
        # # fig2.show()
        #
        # fig3 = self._plot_line_agg_grp(agg_time, load_incl_chg_df, ZONAL_TOTAL_LOAD)
        # fig3.update_traces(showlegend=True, line_color="red")
        # # fig3.show()
        #
        # subfig = make_subplots()
        # subfig.add_traces(fig.data + fig2.data + fig3.data)
        #
        # subfig.update_layout(title=title_note, barmode="stack")
        # # subfig.show()
        #
        # return subfig

    def get_yearhour_dispatch(self, modeled_year: int, zone: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        # TODO: Implement this method with correct energy logic
        # Safeguad against using the function before it is tested and complete
        raise NotImplementedError("year-hour dispatch aggregation not implemented yet")
        #
        # agg_time = self.YEAR_HOUR
        # power_provided_df, zonal_imports_df, load_incl_chg_df = self.get_load_resource_balance(
        #     zone, modeled_year, agg_time
        # )
        # net_imports_renamed = zonal_imports_df.rename(columns={ZONAL_NET_IMPORTS: POWER_OUT_MW}).assign(
        #     **{POWER_IN_MW: 0},
        #     **{self.AGN_COL: NET_IMPORTS},
        # )
        # power_provided_w_imports = pd.concat([power_provided_df, net_imports_renamed])
        #
        # annual_hourly_dispatch_per_resource_df = power_provided_w_imports[
        #     [agg_time, self.AGN_COL, POWER_OUT_MW]
        # ].set_index(agg_time)
        #
        # return power_provided_df, load_incl_chg_df, annual_hourly_dispatch_per_resource_df

    def get_txpath_hourly(
        self, model_years: list[int], zone: str, agg_time: Literal["chrono_timestamp", "year_hour"]
    ) -> pd.DataFrame:
        # TODO: Figure out robust accounting of TxPath-specific hourly results (implement get_txpath_hourly method)
        raise NotImplementedError("get_txpath_hourly not implemented yet")


######################
### TESTING METHOD ###
######################
def resolve_results_viewer_debug_jig(runme: bool = True) -> dict[str, any]:
    """Debug test fixture for use with ResolveResultsViewer class."""

    # For debugging larger data cases and stepping through the code without pytest overhead, use this test fixture; feel free to edit.
    # For stepping through simple cases and small data sizes, it may be helpful to use pytests in test_run_opt.py (instead of this function).

    if runme is False:
        return {}

    results = {}

    # 1. Prepare inputs for ResolveResultsViewer that refer to local file paths

    start_date = datetime.date(2019, 7, 1)
    end_date = datetime.date(2019, 7, 7)
    DATE_LIST = [start_date + datetime.timedelta(days=x) for x in range((end_date - start_date).days + 1)]

    cases_dict = {
        "case1": {
            "init": {
                "case_results_folder": r"C:\Users\okelly\Documents\GitHub\kit\reports\resolve\Multi Zone, Multi Year Refactor\2025-03-20-15-51-14",
                "chrono_date_list": DATE_LIST,
                "modeled_years": [2030],
                "aggregation_config_df": None,
                "enable_column_filter_on_load": True,
                "additional_columns": None,
                "enable_time_filter_on_load": True,
                "max_files": 10000,
                "emit_dw_weights": False,
            },
            "title_note": "allow_inter_period_dynamics: True",
        },
        # # Add more test cases here
        # case2: {
        #     "init": {}
        # },
    }

    # Loop through the cases and call the ResolveResultsViewer methods
    for case_id in cases_dict:

        # 2. Instantiate the ResolveResultsViewer object
        rv = ResolveHourlyResultsViewer(**cases_dict[case_id]["init"])

        # 3. Call one of the data extraction or aggregation methods
        zone = "CAISO"
        modeled_years = rv.modeled_years
        power_provided_df, zonal_imports_df, load_incl_chg_df = rv.get_load_resource_balance(
            zone, rv.modeled_years, CHRONO_TIMESTAMP
        )

        # 4. Call one of the plotting methods
        fig = rv.create_chrono_soc_plot(
            power_provided_df,
            zonal_imports_df,
            load_incl_chg_df,
            model_year=modeled_years[0],
            zone=zone,
            title_note=f"Dispatch {modeled_years[0]=}<br>" + cases_dict[case_id]["title_note"],
        )
        fig.show()

        # 5. Store the results
        results[case_id] = (fig, power_provided_df, zonal_imports_df, load_incl_chg_df)

    return results


if __name__ == "__main__":
    result = resolve_results_viewer_debug_jig(runme=False)
